import importlib.util
import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, r'c:\Users\SELVAM\OneDrive\Desktop\POS\Project')

from modules.menu import report_exports


OPENPYXL_AVAILABLE = importlib.util.find_spec('openpyxl') is not None


DETAIL_SAMPLE = {
    'header': {
        'period_from': '2026-04-01T00:00:00',
        'period_to': '2026-04-29T23:59:59',
        'generated_at': '2026-04-29T12:44:00',
        'generated_by': 'admin',
    },
    'sales_summary': {
        'paid_receipt_count': 2,
        'gross_sales': 89.5,
        'less_refund_outflow': 3.0,
        'less_vendor_outflow': 1.5,
        'net_after_outflows': 85.0,
    },
    'payment_breakdown': [
        {'method': 'CASH', 'amount': 50.0},
        {'method': 'CARD', 'amount': 39.5},
    ],
    'categories': [
        {
            'category_name': 'Dairy',
            'category_total': 13.0,
            'products': [
                {'product_name': 'Milk', 'unit': 'Each', 'qty_sold': 2, 'line_sales': 13.0},
            ],
        }
    ],
    'top_products': [
        {'rank': 1, 'product_name': 'Milk', 'qty_sold': 2, 'unit': 'Each', 'line_sales': 13.0},
    ],
    'cash_outflows': [
        {'outflow_type': 'REFUND_OUT', 'created_at': '2026-04-29T10:00:00', 'cashier': 'admin', 'amount': 3.0, 'note': 'test'},
    ],
    'excluded': {
        'unpaid_receipts_count': 1,
        'unpaid_receipts_total': 20.0,
        'cancelled_receipts_count': 0,
        'cancelled_receipts_total': 0.0,
    },
}


MINIMAL_DETAIL_SAMPLE = dict(DETAIL_SAMPLE)
MINIMAL_DETAIL_SAMPLE['detail_variant'] = 'minimal'


SUMMARY_SAMPLE = {
    'header': {
        'period_from': '2026-04-01T00:00:00',
        'period_to': '2026-04-29T23:59:59',
        'generated_at': '2026-04-29T12:44:00',
        'generated_by': 'admin',
    },
    'sales_summary': {
        'paid_receipt_count': 2.0,
        'gross_sales': 89.5,
        'less_refund_outflow': 3.0,
        'less_vendor_outflow': 1.5,
        'net_after_outflows': 85.0,
    },
    'sales_by_hour': [
        {'hour_slot': '09:00 - 10:00', 'sales_amount': 42.5},
    ],
    'peak_hour': {'hour_slot': '09:00 - 10:00', 'sales_amount': 42.5},
    'top_products_by_qty_hour': [
        {'hour_slot': '09:00 - 10:00', 'products': [{'rank': 1, 'product_name': 'Milk', 'qty_sold': 2, 'unit': 'Each', 'line_sales': 13.0}]},
    ],
    'top_products_by_sales_hour': [
        {'hour_slot': '09:00 - 10:00', 'products': [{'rank': 1, 'product_name': 'Milk', 'qty_sold': 2, 'unit': 'Each', 'line_sales': 13.0}]},
    ],
    'top_products_by_qty_day': [
        {'rank': 1, 'product_name': 'Milk', 'qty_sold': 2, 'unit': 'Each', 'line_sales': 13.0},
    ],
    'top_products_by_sales_day': [
        {'rank': 1, 'product_name': 'Milk', 'qty_sold': 2, 'unit': 'Each', 'line_sales': 13.0},
    ],
    'excluded': {
        'unpaid_receipts_count': 1,
        'unpaid_receipts_total': 20.0,
        'cancelled_receipts_count': 0,
        'cancelled_receipts_total': 0.0,
    },
}


INACTIVITY_SAMPLE = {
    'header': {
        'period_checked': '2026-04-29T00:00:00',
        'generated_at': '2026-04-29T12:44:00',
        'generated_by': 'admin',
    },
    'sections': [
        {'bucket': 'never', 'title': 'NEVER SOLD', 'products': [{'product_code': 'P01', 'product_name': 'Milk', 'category': 'Dairy', 'last_sold': None}]},
    ],
    'summary': {
        'bucket_counts': {'3_6': 0, '6_12': 0, '1_plus': 0, 'never': 1},
        'total_inactive_products': 1,
    },
}


class ReportExportsTest(unittest.TestCase):
    def test_build_report_filename_templates(self):
        ts = '11apr2026_12-44'
        self.assertEqual(
            report_exports.build_report_filename('detail', 'pdf', timestamp=ts),
            'Sales_record_pdf_11apr2026_12-44.pdf',
        )
        self.assertEqual(
            report_exports.build_report_filename('detail', 'xlsx', timestamp=ts),
            'Sales_record_xlsx_11apr2026_12-44.xlsx',
        )
        self.assertEqual(
            report_exports.build_report_filename('summary', 'pdf', timestamp=ts),
            'Sales_trends_pdf_11apr2026_12-44.pdf',
        )
        self.assertEqual(
            report_exports.build_report_filename('inactivity', 'xlsx', timestamp=ts),
            'Inactivity_report_xlsx_11apr2026_12-44.xlsx',
        )
        self.assertEqual(
            report_exports.build_report_filename('chart', 'pdf', timestamp=ts),
            'Charts_pdf_11apr2026_12-44.pdf',
        )
        with self.assertRaises(ValueError):
            report_exports.build_report_filename('chart', 'xlsx', timestamp=ts)

    def test_pdf_text_helper_uses_report_formatter(self):
        text = report_exports._report_text_for_pdf('detail', DETAIL_SAMPLE)
        self.assertIn('Sales Record & Totals', text)
        self.assertIn('Gross Sales', text)

    def test_pdf_text_helper_uses_minimal_detail_variant(self):
        text = report_exports._report_text_for_pdf('detail', MINIMAL_DETAIL_SAMPLE)
        self.assertIn('4. Cash Outflows Detail', text)
        self.assertIn('5. Other Activity (Unpaid & Cancelled)', text)
        self.assertNotIn('Earnings Broken Down by Category', text)
        self.assertNotIn('Top 10 Best Sellers (By Earnings)', text)

    def test_save_report_pdf_inactivity_rejects_oversized_reports(self):
        threshold = int(getattr(report_exports, 'PDF_RENDER_UNIT_THRESHOLD', 18000))
        oversized = threshold + 1
        large_report = {
            'header': {
                'period_checked': '2026-04-29T00:00:00',
                'generated_at': '2026-04-29T12:44:00',
                'generated_by': 'admin',
            },
            'sections': [
                {
                    'bucket': 'never',
                    'title': 'NEVER SOLD',
                    'products': [
                        {
                            'product_code': f'P{index:05d}',
                            'product_name': f'Product {index}',
                            'category': 'Cat',
                            'last_sold': None,
                        }
                        for index in range(oversized)
                    ],
                }
            ],
            'summary': {
                'bucket_counts': {'3_6': 0, '6_12': 0, '1_plus': 0, 'never': oversized},
                'total_inactive_products': oversized,
            },
        }

        with self.assertRaises(RuntimeError):
            report_exports.save_report_pdf('inactivity', report_data=large_report, out_dir=tempfile.gettempdir(), filename='large.pdf')

    def test_save_report_pdf_detail_rejects_oversized_reports(self):
        threshold = int(getattr(report_exports, 'PDF_RENDER_UNIT_THRESHOLD', 18000))
        oversized = threshold + 1
        large_report = dict(DETAIL_SAMPLE)
        large_report['categories'] = [
            {
                'category_name': 'Bulk',
                'category_total': 0.0,
                'products': [
                    {'product_name': f'Item {index}', 'unit': 'Each', 'qty_sold': 1, 'line_sales': 1.0}
                    for index in range(oversized)
                ],
            }
        ]

        with self.assertRaises(RuntimeError):
            report_exports.save_report_pdf('detail', report_data=large_report, out_dir=tempfile.gettempdir(), filename='large_detail.pdf')

    def test_save_report_pdf_summary_rejects_oversized_reports(self):
        threshold = int(getattr(report_exports, 'PDF_RENDER_UNIT_THRESHOLD', 18000))
        oversized = threshold + 1
        large_report = dict(SUMMARY_SAMPLE)
        large_report['sales_by_hour'] = [
            {'hour_slot': f'{index:02d}:00 - {index:02d}:59', 'sales_amount': 1.0}
            for index in range(oversized)
        ]

        with self.assertRaises(RuntimeError):
            report_exports.save_report_pdf('summary', report_data=large_report, out_dir=tempfile.gettempdir(), filename='large_summary.pdf')

    def test_validate_pdf_export_reports_blocked_reason(self):
        threshold = int(getattr(report_exports, 'PDF_RENDER_UNIT_THRESHOLD', 18000))
        oversized = threshold + 1
        large_report = dict(SUMMARY_SAMPLE)
        large_report['sales_by_hour'] = [
            {'hour_slot': f'{index:02d}:00 - {index:02d}:59', 'sales_amount': 1.0}
            for index in range(oversized)
        ]

        allowed, message, units = report_exports.validate_pdf_export('summary', large_report)

        self.assertFalse(allowed)
        self.assertIsNotNone(message)
        self.assertGreater(units, threshold)
        self.assertIn('render safely', message)

    def test_save_report_pdf_chart_creates_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            out_path = report_exports.save_report_pdf(
                'chart',
                report_data=SUMMARY_SAMPLE,
                out_dir=temp_dir,
                filename='chart.pdf',
            )

            self.assertTrue(Path(out_path).exists())

    @unittest.skipUnless(OPENPYXL_AVAILABLE, 'openpyxl is required for XLSX export tests')
    def test_save_report_xlsx_detail_creates_sheet_layout(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            out_path = report_exports.save_report_xlsx(
                'detail',
                report_data=DETAIL_SAMPLE,
                out_dir=temp_dir,
                filename='detail.xlsx',
            )

            self.assertTrue(Path(out_path).exists())

            from openpyxl import load_workbook

            workbook = load_workbook(out_path)
            self.assertEqual(workbook.sheetnames, ['Summary', 'Payments', 'Categories', 'Top Products', 'Outflows', 'Excluded'])
            self.assertEqual(workbook['Summary']['A1'].value, 'Field')
            self.assertEqual(workbook['Payments']['A2'].value, 'CASH')
            self.assertEqual(workbook['Categories']['A2'].value, 'Dairy')
            self.assertEqual(workbook['Top Products']['B2'].value, 'Milk')

    @unittest.skipUnless(OPENPYXL_AVAILABLE, 'openpyxl is required for XLSX export tests')
    def test_save_report_xlsx_detail_minimal_omits_breakdown_sheets(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            out_path = report_exports.save_report_xlsx(
                'detail',
                report_data=MINIMAL_DETAIL_SAMPLE,
                out_dir=temp_dir,
                filename='detail_minimal.xlsx',
            )

            self.assertTrue(Path(out_path).exists())

            from openpyxl import load_workbook

            workbook = load_workbook(out_path)
            self.assertEqual(workbook.sheetnames, ['Summary', 'Payments', 'Outflows', 'Excluded'])
            self.assertNotIn('Categories', workbook.sheetnames)
            self.assertNotIn('Top Products', workbook.sheetnames)

    @unittest.skipUnless(OPENPYXL_AVAILABLE, 'openpyxl is required for XLSX export tests')
    def test_save_report_xlsx_inactivity_creates_bucket_sheets(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            out_path = report_exports.save_report_xlsx(
                'inactivity',
                report_data=INACTIVITY_SAMPLE,
                out_dir=temp_dir,
                filename='inactivity.xlsx',
            )

            self.assertTrue(Path(out_path).exists())

            from openpyxl import load_workbook

            workbook = load_workbook(out_path)
            self.assertEqual(workbook.sheetnames, ['Summary', 'Bucket 3-6', 'Bucket 6-12', 'Bucket 1 Plus', 'Never Sold'])
            self.assertEqual(workbook['Never Sold']['A2'].value, 'P01')
            self.assertEqual(workbook['Summary']['A2'].value, 'Period Checked')

    @unittest.skipUnless(OPENPYXL_AVAILABLE, 'openpyxl is required for XLSX export tests')
    def test_save_report_xlsx_summary_creates_hour_sheets(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            out_path = report_exports.save_report_xlsx(
                'summary',
                report_data=SUMMARY_SAMPLE,
                out_dir=temp_dir,
                filename='summary.xlsx',
            )

            self.assertTrue(Path(out_path).exists())

            from openpyxl import load_workbook

            workbook = load_workbook(out_path)
            self.assertEqual(workbook.sheetnames, [
                'Summary',
                'Sales By Hour',
                'Top Sales By Hour - Pieces',
                'Top Sales By Hour - Weight',
                'Top Qty By Hour - Pieces',
                'Top Qty By Hour - Weight',
                'Top Sales By Day - Pieces',
                'Top Sales By Day - Weight',
                'Top Qty By Day - Pieces',
                'Top Qty By Day - Weight',
                'Excluded',
            ])
            self.assertEqual(workbook['Sales By Hour']['A2'].value, '09:00 - 10:00')
            self.assertEqual(workbook['Top Qty By Day - Pieces']['B2'].value, 'Milk')


if __name__ == '__main__':
    unittest.main()
