"""Report export helpers for PDF and XLSX output."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from PyQt5.QtGui import QFont, QTextDocument
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtWidgets import QApplication

from modules.menu import report_viewers


EXPORT_ROOT = Path.home() / 'POS_Exports' / 'Reports'
_REPORT_STEMS = {
    'detail': 'Sales_record',
    'summary': 'Sales_trends',
    'chart': 'Charts',
    'inactivity': 'Inactivity_report',
}
PDF_RENDER_UNIT_THRESHOLD = 5000
# Backward-compatible alias.
INACTIVITY_PDF_ROW_THRESHOLD = PDF_RENDER_UNIT_THRESHOLD
PDF_TOO_LARGE_MESSAGE = 'PDF export skipped: the selected report is too large to render safely.'


def _normalize_report_type(report_type: Any) -> str:
    rpt = str(report_type or 'summary').strip().lower()
    return rpt if rpt in _REPORT_STEMS else 'summary'


def _timestamp_for_filename(now: datetime | None = None) -> str:
    return (now or datetime.now()).strftime('%d%b%Y_%H-%M').lower()


def _report_stem(report_type: Any) -> str:
    return _REPORT_STEMS.get(_normalize_report_type(report_type), 'Sales_trends')


def build_report_filename(report_type: Any, fmt: str, *, timestamp: str | None = None) -> str:
    fmt_normalized = str(fmt or '').strip().lower()
    if fmt_normalized not in {'pdf', 'xlsx'}:
        raise ValueError(f'Unsupported export format: {fmt}')

    rpt = _normalize_report_type(report_type)
    if rpt == 'chart' and fmt_normalized == 'xlsx':
        raise ValueError('Chart reports can only be exported to PDF')

    ts = str(timestamp or _timestamp_for_filename())
    return f"{_report_stem(rpt)}_{fmt_normalized}_{ts}.{fmt_normalized}"


def _ensure_exports_folder(out_dir: Path | str | None = None) -> Path:
    folder = Path(out_dir) if out_dir is not None else EXPORT_ROOT
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def _safe_text_lines(value: Any) -> list[str]:
    if isinstance(value, str):
        return value.splitlines() or ['']
    if isinstance(value, Iterable):
        return [str(item) for item in value]
    return [str(value)]


def _report_text_for_pdf(report_type: Any, report_data: dict | None) -> str:
    rpt = _normalize_report_type(report_type)
    payload = report_data or {}

    if rpt == 'detail':
        text, _, _, _ = report_viewers._format_detailed_report_text(payload)
        return text
    if rpt == 'summary':
        text, _, _, _ = report_viewers._format_summary_report_text(payload)
        return text
    if rpt == 'inactivity':
        text, _, _, _ = report_viewers._format_inactivity_report_text(payload)
        return text
    raise ValueError(f'Unsupported PDF report type: {report_type}')


def _count_nested_items(value: Any) -> int:
    if isinstance(value, dict):
        total = 0
        for item in value.values():
            total += _count_nested_items(item)
        return total or len(value)
    if isinstance(value, (list, tuple, set)):
        total = 0
        for item in value:
            total += _count_nested_items(item)
        return total or len(value)
    return 1


def estimate_report_render_units(report_type: Any, report_data: dict | None) -> int:
    """Estimate report size from raw data before PDF rendering."""
    rpt = _normalize_report_type(report_type)
    payload = report_data or {}

    if rpt == 'detail':
        sales = payload.get('sales_summary') or {}
        payments = payload.get('payment_breakdown') or []
        categories = payload.get('categories') or []
        top_products = payload.get('top_products') or []
        outflows = payload.get('cash_outflows') or []
        excluded = payload.get('excluded') or {}
        detail_variant = report_viewers._detail_report_variant(payload)
        include_breakdowns = detail_variant != 'minimal'
        category_products = sum(len(category.get('products') or []) for category in categories if isinstance(category, dict)) if include_breakdowns else 0
        return (
            len(payments)
            + (len(categories) if include_breakdowns else 0)
            + category_products
            + (len(top_products) if include_breakdowns else 0)
            + len(outflows)
            + len(excluded)
            + len(sales)
            + len(payload.get('header') or {})
        )

    if rpt == 'summary':
        sales = payload.get('sales_summary') or {}
        sales_by_hour = payload.get('sales_by_hour') or []
        peak_hour = payload.get('peak_hour') or {}
        top_qty_by_hour = payload.get('top_products_by_qty_hour') or []
        top_sales_by_hour = payload.get('top_products_by_sales_hour') or []
        top_qty_day = payload.get('top_products_by_qty_day') or []
        top_sales_day = payload.get('top_products_by_sales_day') or []
        excluded = payload.get('excluded') or {}
        grouped_products = sum(
            len(group.get('products') or [])
            for group in list(top_qty_by_hour) + list(top_sales_by_hour)
            if isinstance(group, dict)
        )
        return (
            len(sales_by_hour)
            + len(top_qty_by_hour)
            + len(top_sales_by_hour)
            + grouped_products
            + len(top_qty_day)
            + len(top_sales_day)
            + len(excluded)
            + len(sales)
            + len(peak_hour)
            + len(payload.get('header') or {})
        )

    if rpt == 'inactivity':
        header = payload.get('header') or {}
        sections = payload.get('sections') or []
        summary = payload.get('summary') or {}
        section_products = sum(len(section.get('products') or []) for section in sections if isinstance(section, dict))
        return len(sections) + section_products + len(summary) + len(header)

    return _count_nested_items(payload)


def estimate_inactivity_report_rows(report_data: dict | None) -> int:
    """Compatibility wrapper for inactivity report size estimate."""
    return estimate_report_render_units('inactivity', report_data)


def validate_pdf_export(report_type: Any, report_data: dict | None, *, threshold: int | None = None) -> tuple[bool, str | None]:
    """Check whether a report is safe to render to PDF.

    Returns:
        (is_allowed, message_if_blocked)
    """
    rpt = _normalize_report_type(report_type)
    estimated_units = estimate_report_render_units(rpt, report_data)
    limit = PDF_RENDER_UNIT_THRESHOLD if threshold is None else int(threshold)
    if estimated_units > limit:
        return False, PDF_TOO_LARGE_MESSAGE
    return True, None


def save_report_pdf(
    report_type: Any,
    *,
    report_data: dict | None = None,
    out_dir: Path | str | None = None,
    filename: str | None = None,
) -> Path:
    rpt = _normalize_report_type(report_type)
    folder = _ensure_exports_folder(out_dir)
    file_name = filename or build_report_filename(rpt, 'pdf')
    out_path = folder / file_name

    # Guard against rendering very large reports — check size and raise for callers.
    estimated_units = estimate_report_render_units(rpt, report_data)
    limit = PDF_RENDER_UNIT_THRESHOLD
    if estimated_units > limit:
        raise RuntimeError(PDF_TOO_LARGE_MESSAGE)

    if QApplication.instance() is None:
        _qt_app = QApplication([])

    if rpt == 'chart':
        from modules.menu import report_charts

        return report_charts.save_chart_report_pdf(report_data or {}, out_path)

    text = _report_text_for_pdf(rpt, report_data)

    document = QTextDocument()
    document.setDefaultFont(QFont('Courier New', 10))
    document.setPlainText(text)

    printer = QPrinter(QPrinter.HighResolution)
    printer.setOutputFormat(QPrinter.PdfFormat)
    printer.setOutputFileName(str(out_path))
    document.print_(printer)
    return out_path


def _write_header_row(sheet, headers: list[str]) -> None:
    # Write headers explicitly so subsequent rows can be written cell-by-cell
    # (which allows us to set number formats per-cell).
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_index, value=header)


def _auto_fit_columns(sheet) -> None:
    widths: dict[str, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = '' if cell.value is None else str(cell.value)
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), min(len(value), 60))

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = max(10, min(width + 2, 60))


def _style_table_sheet(sheet) -> None:
    try:
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return

    if sheet.max_row >= 1:
        header_fill = PatternFill('solid', fgColor='D9E2F3')
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill


def _add_table_sheet(workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    _write_header_row(sheet, headers)
    # Write rows cell-by-cell to allow number formatting.
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            # Apply 2-decimal format for floats (figures) to match PDF formatting.
            try:
                if isinstance(value, float):
                    cell.number_format = '#,##0.00'
            except Exception:
                pass
    _style_table_sheet(sheet)
    _auto_fit_columns(sheet)


def _add_key_value_sheet(workbook, title: str, rows: list[tuple[Any, Any]]) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.cell(row=1, column=1, value='Field')
    sheet.cell(row=1, column=2, value='Value')
    for r_idx, (key, value) in enumerate(rows, start=2):
        sheet.cell(row=r_idx, column=1, value=key)
        cell = sheet.cell(row=r_idx, column=2, value=value)
        try:
            if isinstance(value, float):
                cell.number_format = '#,##0.00'
        except Exception:
            pass
    _style_table_sheet(sheet)
    _auto_fit_columns(sheet)


def _detail_workbook_data(report_data: dict) -> dict[str, list[list[Any]] | list[tuple[Any, Any]]]:
    header = report_data.get('header') or {}
    sales = report_data.get('sales_summary') or {}
    payments = report_data.get('payment_breakdown') or []
    categories = report_data.get('categories') or []
    top_products = report_data.get('top_products') or []
    outflows = report_data.get('cash_outflows') or []
    excluded = report_data.get('excluded') or {}
    detail_variant = report_viewers._detail_report_variant(report_data)
    include_breakdowns = detail_variant != 'minimal'

    summary_rows: list[tuple[Any, Any]] = [
        ('Period From', header.get('period_from') or '-'),
        ('Period To', header.get('period_to') or '-'),
        ('Generated At', header.get('generated_at') or '-'),
        ('Generated By', header.get('generated_by') or '-'),
        ('Paid Receipt Count', sales.get('paid_receipt_count') or 0),
        ('Gross Sales', sales.get('gross_sales') or 0),
        ('Refund Outflow', sales.get('less_refund_outflow') or 0),
        ('Vendor Outflow', sales.get('less_vendor_outflow') or 0),
        ('Net After Outflows', sales.get('net_after_outflows') or 0),
    ]

    payment_rows = [[row.get('method') or '', row.get('amount') or 0] for row in payments]

    category_rows: list[list[Any]] = []
    if include_breakdowns:
        for category in categories:
            cat_name = category.get('category_name') or 'Uncategorized'
            for product in category.get('products') or []:
                category_rows.append([
                    cat_name,
                    product.get('product_name') or '',
                    product.get('unit') or '',
                    product.get('qty_sold') or 0,
                    product.get('line_sales') or 0,
                ])

    top_rows = []
    if include_breakdowns:
        top_rows = [[row.get('rank') or '', row.get('product_name') or '', row.get('qty_sold') or 0, row.get('unit') or '', row.get('line_sales') or 0] for row in top_products]

    outflow_rows = [[row.get('outflow_type') or '', row.get('created_at') or '', row.get('cashier') or '', row.get('amount') or 0, row.get('note') or ''] for row in outflows]

    excluded_rows = [
        ('Unpaid Receipts Count', excluded.get('unpaid_receipts_count') or 0),
        ('Unpaid Receipts Total', excluded.get('unpaid_receipts_total') or 0),
        ('Cancelled Receipts Count', excluded.get('cancelled_receipts_count') or 0),
        ('Cancelled Receipts Total', excluded.get('cancelled_receipts_total') or 0),
    ]

    return {
        'detail_variant': detail_variant,
        'summary_rows': summary_rows,
        'payment_rows': payment_rows,
        'category_rows': category_rows,
        'top_rows': top_rows,
        'outflow_rows': outflow_rows,
        'excluded_rows': excluded_rows,
    }


def _summary_workbook_data(report_data: dict) -> dict[str, list[list[Any]] | list[tuple[Any, Any]]]:
    header = report_data.get('header') or {}
    sales = report_data.get('sales_summary') or {}
    sales_by_hour = report_data.get('sales_by_hour') or []
    peak_hour = report_data.get('peak_hour') or {}
    top_qty_by_hour = report_data.get('top_products_by_qty_hour') or []
    top_sales_by_hour = report_data.get('top_products_by_sales_hour') or []
    top_qty_day = report_data.get('top_products_by_qty_day') or []
    top_sales_day = report_data.get('top_products_by_sales_day') or []
    excluded = report_data.get('excluded') or {}

    summary_rows: list[tuple[Any, Any]] = [
        ('Period From', header.get('period_from') or '-'),
        ('Period To', header.get('period_to') or '-'),
        ('Generated At', header.get('generated_at') or '-'),
        ('Generated By', header.get('generated_by') or '-'),
        ('Avg Paid Receipt Count', sales.get('paid_receipt_count') or 0),
        ('Avg Gross Sales', sales.get('gross_sales') or 0),
        ('Avg Refund Outflow', sales.get('less_refund_outflow') or 0),
        ('Avg Vendor Outflow', sales.get('less_vendor_outflow') or 0),
        ('Avg Net After Outflows', sales.get('net_after_outflows') or 0),
        ('Peak Avg Hour', peak_hour.get('hour_slot') or '-'),
    ]

    by_hour_rows = [[row.get('hour_slot') or '', row.get('sales_amount') or 0] for row in sales_by_hour]

    def _split_by_unit_type_export(products: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        """Split products into pieces (each/ea) and weight (kg/g) groups."""
        pieces = []
        weight = []
        for p in products:
            unit_lower = str(p.get('unit', '')).lower()
            if unit_lower in ('each', 'ea'):
                pieces.append(p)
            elif 'kg' in unit_lower or 'g' in unit_lower:
                weight.append(p)
        return pieces, weight

    def _sort_products_export(products: list[dict[str, Any]], primary_metric: str = 'qty') -> list[dict[str, Any]]:
        """Sort products by primary metric desc, secondary by line_sales desc."""
        if primary_metric == 'qty':
            return sorted(products, key=lambda p: (float(p.get('qty_sold', 0)), float(p.get('line_sales', 0))), reverse=True)
        else:
            return sorted(products, key=lambda p: (float(p.get('line_sales', 0)), float(p.get('qty_sold', 0))), reverse=True)

    def _flatten_grouped_split(groups: list[dict[str, Any]], primary_metric: str = 'qty', per_unit_limit: int = 3) -> tuple[list[list[Any]], list[list[Any]]]:
        """Flatten hourly groups split by unit type: returns (pieces_rows, weight_rows)."""
        pieces_rows: list[list[Any]] = []
        weight_rows: list[list[Any]] = []
        
        for group in groups:
            hour_slot = group.get('hour_slot') or ''
            products = group.get('products') or []
            
            pieces, weight = _split_by_unit_type_export(products)
            pieces = _sort_products_export(pieces, primary_metric)[:per_unit_limit]
            weight = _sort_products_export(weight, primary_metric)[:per_unit_limit]
            
            for idx, row in enumerate(pieces, start=1):
                pieces_rows.append([
                    hour_slot,
                    idx,
                    row.get('product_name') or '',
                    row.get('qty_sold') or 0,
                    row.get('unit') or '',
                    row.get('line_sales') or 0,
                ])
            
            for idx, row in enumerate(weight, start=1):
                weight_rows.append([
                    hour_slot,
                    idx,
                    row.get('product_name') or '',
                    row.get('qty_sold') or 0,
                    row.get('unit') or '',
                    row.get('line_sales') or 0,
                ])
        
        return pieces_rows, weight_rows

    def _day_section_split(rows: list[dict[str, Any]], primary_metric: str = 'qty', pieces_limit: int = 10, weight_limit: int = 5) -> tuple[list[list[Any]], list[list[Any]]]:
        """Split daily sellers by unit type: returns (pieces_rows, weight_rows)."""
        pieces, weight = _split_by_unit_type_export(rows)
        pieces = _sort_products_export(pieces, primary_metric)[:pieces_limit]
        weight = _sort_products_export(weight, primary_metric)[:weight_limit]
        
        pieces_rows = [[idx, row.get('product_name') or '', row.get('qty_sold') or 0, row.get('unit') or '', row.get('line_sales') or 0] 
                       for idx, row in enumerate(pieces, start=1)]
        weight_rows = [[idx, row.get('product_name') or '', row.get('qty_sold') or 0, row.get('unit') or '', row.get('line_sales') or 0] 
                       for idx, row in enumerate(weight, start=1)]
        
        return pieces_rows, weight_rows

    # Flatten hourly groups (sections 3 & 4 are now swapped in viewer)
    sales_hour_pieces, sales_hour_weight = _flatten_grouped_split(top_sales_by_hour, primary_metric='sales')
    qty_hour_pieces, qty_hour_weight = _flatten_grouped_split(top_qty_by_hour, primary_metric='qty')
    
    # Day sections (sections 5 & 6 are now swapped in viewer)
    sales_day_pieces, sales_day_weight = _day_section_split(top_sales_day, primary_metric='sales')
    qty_day_pieces, qty_day_weight = _day_section_split(top_qty_day, primary_metric='qty')

    excluded_rows = [
        ('Unpaid Receipts Count', excluded.get('unpaid_receipts_count') or 0),
        ('Unpaid Receipts Total', excluded.get('unpaid_receipts_total') or 0),
        ('Cancelled Receipts Count', excluded.get('cancelled_receipts_count') or 0),
        ('Cancelled Receipts Total', excluded.get('cancelled_receipts_total') or 0),
    ]

    return {
        'summary_rows': summary_rows,
        'sales_by_hour_rows': by_hour_rows,
        'sales_hour_pieces': sales_hour_pieces,
        'sales_hour_weight': sales_hour_weight,
        'qty_hour_pieces': qty_hour_pieces,
        'qty_hour_weight': qty_hour_weight,
        'sales_day_pieces': sales_day_pieces,
        'sales_day_weight': sales_day_weight,
        'qty_day_pieces': qty_day_pieces,
        'qty_day_weight': qty_day_weight,
        'excluded_rows': excluded_rows,
    }


def _inactivity_workbook_data(report_data: dict) -> dict[str, list[list[Any]] | list[tuple[Any, Any]]]:
    header = report_data.get('header') or {}
    sections = report_data.get('sections') or []
    summary = report_data.get('summary') or {}

    summary_rows: list[tuple[Any, Any]] = [
        ('Period Checked', header.get('period_checked') or '-'),
        ('Generated At', header.get('generated_at') or '-'),
        ('Generated By', header.get('generated_by') or '-'),
        ('3-6 Months No Sale', (summary.get('bucket_counts') or {}).get('3_6', 0)),
        ('6-12 Months No Sale', (summary.get('bucket_counts') or {}).get('6_12', 0)),
        ('More Than 1 Year No Sale', (summary.get('bucket_counts') or {}).get('1_plus', 0)),
        ('Never Sold', (summary.get('bucket_counts') or {}).get('never', 0)),
        ('Total Inactive Products', summary.get('total_inactive_products') or 0),
    ]

    buckets: dict[str, list[list[Any]]] = {
        '3_6': [],
        '6_12': [],
        '1_plus': [],
        'never': [],
    }
    for section in sections:
        bucket = str(section.get('bucket') or '')
        rows = buckets.get(bucket)
        if rows is None:
            continue
        for row in section.get('products') or []:
            rows.append([
                row.get('product_code') or '',
                row.get('product_name') or '',
                row.get('category') or '',
                row.get('last_sold') or '',
            ])

    return {
        'summary_rows': summary_rows,
        'bucket_rows': buckets,
    }


def save_report_xlsx(
    report_type: Any,
    *,
    report_data: dict | None = None,
    out_dir: Path | str | None = None,
    filename: str | None = None,
) -> Path:
    rpt = _normalize_report_type(report_type)
    if rpt == 'chart':
        raise ValueError('Chart reports can only be exported to PDF')

    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError('Missing dependency: openpyxl') from exc

    folder = _ensure_exports_folder(out_dir)
    file_name = filename or build_report_filename(rpt, 'xlsx')
    out_path = folder / file_name

    workbook = Workbook()
    workbook.remove(workbook.active)
    payload = report_data or {}

    if rpt == 'detail':
        data = _detail_workbook_data(payload)
        _add_key_value_sheet(workbook, 'Summary', data['summary_rows'])
        _add_table_sheet(workbook, 'Payments', ['Method', 'Amount'], data['payment_rows'])
        if data.get('detail_variant') != 'minimal':
            _add_table_sheet(workbook, 'Categories', ['Category', 'Product Name', 'Unit', 'Qty Sold', 'Line Sales'], data['category_rows'])
            _add_table_sheet(workbook, 'Top Products', ['Rank', 'Product Name', 'Qty Sold', 'Unit', 'Amount'], data['top_rows'])
        _add_table_sheet(workbook, 'Outflows', ['Type', 'Date/Time', 'Cashier', 'Amount', 'Note'], data['outflow_rows'])
        _add_key_value_sheet(workbook, 'Excluded', data['excluded_rows'])
    elif rpt == 'summary':
        data = _summary_workbook_data(payload)
        _add_key_value_sheet(workbook, 'Summary', data['summary_rows'])
        _add_table_sheet(workbook, 'Sales By Hour', ['Hour Slot', 'Avg Revenue'], data['sales_by_hour_rows'])
        _add_table_sheet(workbook, 'Top Sales By Hour - Pieces', ['Hour Slot', 'Rank', 'Product Name', 'Avg Qty', 'Unit', 'Avg Revenue'], data['sales_hour_pieces'])
        _add_table_sheet(workbook, 'Top Sales By Hour - Weight', ['Hour Slot', 'Rank', 'Product Name', 'Avg Qty', 'Unit', 'Avg Revenue'], data['sales_hour_weight'])
        _add_table_sheet(workbook, 'Top Qty By Hour - Pieces', ['Hour Slot', 'Rank', 'Product Name', 'Avg Qty', 'Unit', 'Avg Revenue'], data['qty_hour_pieces'])
        _add_table_sheet(workbook, 'Top Qty By Hour - Weight', ['Hour Slot', 'Rank', 'Product Name', 'Avg Qty', 'Unit', 'Avg Revenue'], data['qty_hour_weight'])
        _add_table_sheet(workbook, 'Top Sales By Day - Pieces', ['Rank', 'Product Name', 'Avg Qty', 'Unit', 'Avg Revenue'], data['sales_day_pieces'])
        _add_table_sheet(workbook, 'Top Sales By Day - Weight', ['Rank', 'Product Name', 'Avg Qty', 'Unit', 'Avg Revenue'], data['sales_day_weight'])
        _add_table_sheet(workbook, 'Top Qty By Day - Pieces', ['Rank', 'Product Name', 'Avg Qty', 'Unit', 'Avg Revenue'], data['qty_day_pieces'])
        _add_table_sheet(workbook, 'Top Qty By Day - Weight', ['Rank', 'Product Name', 'Avg Qty', 'Unit', 'Avg Revenue'], data['qty_day_weight'])
        _add_key_value_sheet(workbook, 'Excluded', data['excluded_rows'])
    else:
        data = _inactivity_workbook_data(payload)
        _add_key_value_sheet(workbook, 'Summary', data['summary_rows'])
        _add_table_sheet(workbook, 'Bucket 3-6', ['Product Code', 'Product Name', 'Category', 'Last Sold'], data['bucket_rows']['3_6'])
        _add_table_sheet(workbook, 'Bucket 6-12', ['Product Code', 'Product Name', 'Category', 'Last Sold'], data['bucket_rows']['6_12'])
        _add_table_sheet(workbook, 'Bucket 1 Plus', ['Product Code', 'Product Name', 'Category', 'Last Sold'], data['bucket_rows']['1_plus'])
        _add_table_sheet(workbook, 'Never Sold', ['Product Code', 'Product Name', 'Category', 'Last Sold'], data['bucket_rows']['never'])

    workbook.save(str(out_path))
    return out_path
